from datetime import date
from datetime import datetime
import flask
import openpyxl
from flask import json, jsonify, request
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font,Alignment,Border,Side
from openpyxl.drawing import drawing


class AppVRG:

    def generate_VRG(self, data):
        req_data = data
        now = date.today()
        full = str(now.month) + "/" + str(now.day) + "/" + str(now.year)
        wb = openpyxl.Workbook()
        wb = load_workbook('New_VRG_template.xlsx')

        wb.template = True
        ws = wb.active
        ws.cell(row=6, column=1).value = 'Last Modified:' + full
        ws.cell(row=6, column=4).value = str(
            ws.cell(row=6, column=4).value) + str(req_data['userName']).title()
        ws.cell(row=6, column=4).alignment = openpyxl.styles.Alignment(
            horizontal='center', vertical='center', wrap_text=True)
        datestring = datetime.strftime(datetime.now(), ' %Y_%m_%d %HH %MM')
        c = 8
        ws.cell(row=6, column=5).value = str(
            ws.cell(row=6, column=5).value) + str(req_data['ProjectName']).title()
        fname = req_data['ProjectName'] + datestring

        for i in range(0, len(req_data['pages'])):
            r = 9
            if 'pageName' in req_data['pages'][i]:
                ws.cell(row=8, column=c).value = str(
                    req_data['pages'][i]['pageName']).upper()
            if 'pageView' in req_data['pages'][i]:
                ws.cell(
                    row=r, column=c).value = req_data['pages'][i]['pageView']

            # site related stuff
            if 'sitebrand' in req_data:
                ws.cell(row=r + 1, column=c).value = req_data['sitebrand']
            if 'siteName' in req_data:
                ws.cell(row=r + 2, column=c).value = req_data['siteName']
            if 'siteType' in req_data:
                ws.cell(row=r + 3, column=c).value = req_data['siteType']
            if 'siteEnviroment' in req_data:
                ws.cell(row=r + 4, column=c).value = req_data['siteEnviroment']
            if 'siteAppVersion' in req_data:
                ws.cell(row=r + 5, column=c).value = req_data['siteAppVersion']
            if 'siteLastBuildDate' in req_data:
                ws.cell(
                    row=r + 6, column=c).value = req_data['siteLastBuildDate']

            if 'pagePath' in req_data['pages'][i]:
                ws.cell(
                    row=r + 7, column=c).value = req_data['pages'][i]['pagePath']
            #if 'pageName' in req_data['pages'][i]:
            #    ws.cell(row=r + 8, column=c).value = req_data['sitebrand'] + "." + req_data['siteName'] + "." + \
            #                                         req_data['pages'][i]['pageName']
            if 'pageName' in req_data['pages'][i]:
                ws.cell(row=r + 8, column=c).value = req_data['pages'][i]['pageName']

            if 'pageUrl' in req_data['pages'][i]:
                ws.cell(
                    row=r + 9, column=c).value = req_data['pages'][i]['pageUrl']
            if 'pageReferrer' in req_data['pages'][i]:
                ws.cell(
                    row=r + 10, column=c).value = req_data['pages'][i]['pageReferrer']
            if 'pageHierarchy' in req_data['pages'][i]:
                ws.cell(
                    row=r + 11, column=c).value = req_data['pages'][i]['pageHierarchy']
            if 'pageLanguage' in req_data['pages'][i]:
                ws.cell(
                    row=r + 12, column=c).value = req_data['pages'][i]['pageLanguage']
            if 'pageAccessibilty' in req_data['pages'][i]:
                ws.cell(
                    row=r + 13, column=c).value = req_data['pages'][i]['pageAccessibilty']

            if 'userAuthState' in req_data['pages'][i]:
                ws.cell(
                    row=r + 14, column=c).value = req_data['pages'][i]['userAuthState']
            if 'userType' in req_data['pages'][i]:
                ws.cell(
                    row=r + 15, column=c).value = req_data['pages'][i]['userType']
            if 'userId' in req_data['pages'][i]:
                ws.cell(
                    row=r + 16, column=c).value = req_data['pages'][i]['userId']

            if 'eventError' in req_data['pages'][i]:
                ws.cell(
                    row=r + 44, column=c).value = req_data['pages'][i]['eventError']
            if 'errorsType' in req_data['pages'][i]:
                ws.cell(
                    row=r + 45, column=c).value = req_data['pages'][i]['errorsType']
            if 'errorsSubType' in req_data['pages'][i]:
                ws.cell(
                    row=r + 46, column=c).value = req_data['pages'][i]['errorsSubType']
            if 'errorsField' in req_data['pages'][i]:
                ws.cell(
                    row=r + 47, column=c).value = req_data['pages'][i]['errorsField']
            if 'errorMessage' in req_data['pages'][i]:
                ws.cell(
                    row=r + 48, column=c).value = req_data['pages'][i]['errorMessage']
            if 'errorsCode' in req_data['pages'][i]:
                ws.cell(
                    row=r + 49, column=c).value = req_data['pages'][i]['errorsCode']
            c = c + 1

        for parent in range(0, len(req_data['forms'])):
            r = 8
            for child in range(0, len(req_data['forms'][parent]['steps'])):
                if 'pageName' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(row=8, column=c).value = str(
                        req_data['forms'][parent]['steps'][child]['pageName']).upper()
                if 'pageView' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 1, column=c).value = req_data['forms'][parent]['steps'][child]['pageView']

                if 'sitebrand' in req_data:
                    ws.cell(row=r + 2, column=c).value = req_data['sitebrand']
                if 'siteName' in req_data:
                    ws.cell(row=r + 3, column=c).value = req_data['siteName']
                if 'siteType' in req_data:
                    ws.cell(row=r + 4, column=c).value = req_data['siteType']
                if 'siteEnviroment' in req_data:
                    ws.cell(
                        row=r + 5, column=c).value = req_data['siteEnviroment']
                if 'siteAppVersion' in req_data:
                    ws.cell(
                        row=r + 6, column=c).value = req_data['siteAppVersion']
                if 'siteLastBuildDate' in req_data:
                    ws.cell(
                        row=r + 7, column=c).value = req_data['siteLastBuildDate']
                if 'pagePath' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 8, column=c).value = req_data['forms'][parent]['steps'][child]['pagePath']
                if 'pageName' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(row=r + 9, column=c).value = req_data['forms'][parent]['steps'][child]['pageName']

                if 'pageUrl' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 10, column=c).value = req_data['forms'][parent]['steps'][child]['pageUrl']
                if 'pageReferrer' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 11, column=c).value = req_data['forms'][parent]['steps'][child]['pageReferrer']
                if 'pageHierarchy' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 12, column=c).value = str(req_data['forms'][parent]['steps'][child]['pageHierarchy'])
                if 'pageLanguage' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 13, column=c).value = req_data['forms'][parent]['steps'][child]['pageLanguage']
                if 'pageAccessibilty' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 14, column=c).value = req_data['forms'][parent]['steps'][child]['pageAccessibilty']

                if 'userAuthState' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 15, column=c).value = req_data['forms'][parent]['steps'][child]['userAuthState']
                if 'userType' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 16, column=c).value = req_data['forms'][parent]['steps'][child]['userType']
                if 'userId' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 17, column=c).value = req_data['forms'][parent]['steps'][child]['userId']

                if 'formView' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 18, column=c).value = req_data['forms'][parent]['steps'][child]['formView']
                if 'formsubmit' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 19, column=c).value = req_data['forms'][parent]['steps'][child]['formsubmit']
                if 'formQualify' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 20, column=c).value = req_data['forms'][parent]['steps'][child]['formQualify']
                if 'formName' in req_data['forms'][parent]:
                    ws.cell(
                        row=r + 21, column=c).value = str(req_data['forms'][parent]['formName']).lower().replace(" ","-")
                if 'uniqueId' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 22, column=c).value = req_data['forms'][parent]['steps'][child]['uniqueId']
                if 'formStep' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 23, column=c).value = req_data['forms'][parent]['steps'][child]['formStep']

                if 'stepName' in req_data['forms'][parent]['steps'][child]:
                    if(req_data['forms'][parent]['steps'][child]['stepName'] != ''):
                        ws.cell(
                            row=r + 24, column=c).value = str(req_data['forms'][parent]['steps'][child]['stepName']).lower().replace(" ","-")
                    else:
                        arr = str(req_data['forms'][parent]['steps'][child]['pageHierarchy']).replace("\\"," ").replace("[","").replace("]","").replace("'","").split(" ")
                        ws.cell(
                            row=r + 24, column=c).value = str(arr[len(arr)-1]).lower().replace(" ","-")

                if 'eventError' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 45, column=c).value = req_data['forms'][parent]['steps'][child]['eventError']
                if 'errorsType' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 46, column=c).value = req_data['forms'][parent]['steps'][child]['errorsType']
                if 'errorsSubType' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 47, column=c).value = req_data['forms'][parent]['steps'][child]['errorsSubType']
                if 'errorsField' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 48, column=c).value = req_data['forms'][parent]['steps'][child]['errorsField']
                if 'errorMessage' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 48, column=c).value = req_data['forms'][parent]['steps'][child]['errorMessage']
                if 'errorsCode' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 50, column=c).value = req_data['forms'][parent]['steps'][child]['errorsCode']

                # Transaction Information in VRG
                if 'transactionId' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 29, column=c).value = req_data['forms'][parent]['steps'][child]['transactionId']
                if 'transactionamount' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 30, column=c).value = req_data['forms'][parent]['steps'][child]['transactionamount']
                if 'transactionServiceFee' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 31, column=c).value = req_data['forms'][parent]['steps'][child]['transactionServiceFee']
                if 'transactionUnit' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 32, column=c).value = req_data['forms'][parent]['steps'][child]['transactionUnit']
                if 'fromtransaction' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 34, column=c).value = req_data['forms'][parent]['steps'][child]['fromtransaction']
                if 'totransaction' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 35, column=c).value = req_data['forms'][parent]['steps'][child]['totransaction']
                if 'isExternal' in req_data['forms'][parent]:
                    ws.cell(
                        row=r + 36, column=c).value = req_data['forms'][parent]['steps'][child]['isExternal']
                if 'istransactionComplete' in req_data['forms'][parent]:
                    ws.cell(
                        row=r + 37, column=c).value = req_data['forms'][parent]['steps'][child]['istransactionComplete']

                # Product Specific Details
                if 'productId' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 86, column=c).value = req_data['forms'][parent]['steps'][child]['productId']
                if 'productPositioning' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 88, column=c).value = req_data['forms'][parent]['steps'][child]['productPositioning']
                if 'productGrouping' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 89, column=c).value = req_data['forms'][parent]['steps'][child]['productGrouping']
                if 'parentProduct' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 90, column=c).value = req_data['forms'][parent]['steps'][child]['parentProduct']
                if 'fulfillment' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 91, column=c).value = req_data['forms'][parent]['steps'][child]['fulfillment']
                if 'adjudication' in req_data['forms'][parent]:
                    ws.cell(
                        row=r + 92, column=c).value = req_data['forms'][parent]['steps'][child]['adjudication']
                if 'productRecommendation' in req_data['forms'][parent]:
                    ws.cell(
                        row=r + 94, column=c).value = req_data['forms'][parent]['steps'][child]['productRecommendation']
                if 'isPaperless' in req_data['forms'][parent]:
                    ws.cell(
                        row=r + 95, column=c).value = req_data['forms'][parent]['steps'][child]['isPaperless']

                # Error Information in VRG
                if 'eventError' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 45, column=c).value = req_data['forms'][parent]['steps'][child]['eventError']
                if 'errorsType' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 46, column=c).value = req_data['forms'][parent]['steps'][child]['errorsType']
                if 'errorsSubType' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 47, column=c).value = req_data['forms'][parent]['steps'][child]['errorsSubType']
                if 'errorsField' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 48, column=c).value = req_data['forms'][parent]['steps'][child]['errorsField']
                if 'errorMessage' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 49, column=c).value = req_data['forms'][parent]['steps'][child]['errorMessage']
                if 'errorsCode' in req_data['forms'][parent]['steps'][child]:
                    ws.cell(
                        row=r + 50, column=c).value = req_data['forms'][parent]['steps'][child]['errorsCode']
                c = c + 1

        for i in range(0, len(req_data['interaction'])):
            r = 9
            if 'pageName' in req_data['interaction'][i]:
                ws.cell(
                    row=8, column=c).value = req_data['interaction'][i]['interactionName']
            if 'pageView' in req_data['interaction'][i]:
                ws.cell(
                    row=r, column=c).value = req_data['interaction'][i]['pageView']

            # site related stuff
            if 'sitebrand' in req_data:
                ws.cell(row=r + 1, column=c).value = req_data['sitebrand']
            if 'siteName' in req_data:
                ws.cell(row=r + 2, column=c).value = req_data['siteName']
            if 'siteType' in req_data:
                ws.cell(row=r + 3, column=c).value = req_data['siteType']
            if 'siteEnviroment' in req_data:
                ws.cell(row=r + 4, column=c).value = req_data['siteEnviroment']
            if 'siteAppVersion' in req_data:
                ws.cell(row=r + 5, column=c).value = req_data['siteAppVersion']
            if 'siteLastBuildDate' in req_data:
                ws.cell(
                    row=r + 6, column=c).value = req_data['siteLastBuildDate']

            if 'pagePath' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 7, column=c).value = req_data['interaction'][i]['pagePath']
            if 'pageName' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 8, column=c).value = req_data['interaction'][i]['pageName']
            if 'pageUrl' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 9, column=c).value = req_data['interaction'][i]['pageUrl']
            if 'pageReferrer' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 10, column=c).value = req_data['interaction'][i]['pageReferrer']
            if 'pageHierarchy' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 11, column=c).value = req_data['interaction'][i]['pageHierarchy']
            if 'pageLanguage' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 12, column=c).value = req_data['interaction'][i]['pageLanguage']
            if 'pageAccessibilty' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 13, column=c).value = req_data['interaction'][i]['pageAccessibilty']

            if 'userAuthState' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 14, column=c).value = req_data['interaction'][i]['userAuthState']
            if 'userType' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 15, column=c).value = req_data['interaction'][i]['userType']
            if 'userid' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 16, column=c).value = req_data['interaction'][i]['userid']

            if 'eventError' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 44, column=c).value = req_data['interaction'][i]['eventError']

            if 'siteInteractionEvent' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 55, column=c).value = req_data['interaction'][i]['siteInteractionEvent']
            if 'interactionName' in req_data['interaction'][i]:
                ws.cell(
                    row=r + 56, column=c).value = req_data['interaction'][i]['interactionName']
            c = c + 1

        for i in range(0, len(req_data['download'])):
            r = 9
            if 'pageName' in req_data['download'][i]:
                ws.cell(
                    row=8, column=c).value = req_data['download'][i]['download_filename']
            if 'pageView' in req_data['download'][i]:
                ws.cell(
                    row=r, column=c).value = req_data['download'][i]['pageView']

            # site related stuff
            if 'sitebrand' in req_data:
                ws.cell(row=r + 1, column=c).value = req_data['sitebrand']
            if 'siteName' in req_data:
                ws.cell(row=r + 2, column=c).value = req_data['siteName']
            if 'siteType' in req_data:
                ws.cell(row=r + 3, column=c).value = req_data['siteType']
            if 'siteEnviroment' in req_data:
                ws.cell(row=r + 4, column=c).value = req_data['siteEnviroment']
            if 'siteAppVersion' in req_data:
                ws.cell(row=r + 5, column=c).value = req_data['siteAppVersion']
            if 'siteLastBuildDate' in req_data:
                ws.cell(
                    row=r + 6, column=c).value = req_data['siteLastBuildDate']

            if 'pagePath' in req_data['download'][i]:
                ws.cell(
                    row=r + 7, column=c).value = req_data['download'][i]['pagePath']
            if 'pageName' in req_data['download'][i]:
                ws.cell(
                    row=r + 8, column=c).value = req_data['download'][i]['pageName']
            if 'pageUrl' in req_data['download'][i]:
                ws.cell(
                    row=r + 9, column=c).value = req_data['download'][i]['pageUrl']
            if 'pageReferrer' in req_data['download'][i]:
                ws.cell(
                    row=r + 10, column=c).value = req_data['download'][i]['pageReferrer']
            if 'pageHierarchy' in req_data['download'][i]:
                ws.cell(
                    row=r + 11, column=c).value = req_data['download'][i]['pageHierarchy']
            if 'pageLanguage' in req_data['download'][i]:
                ws.cell(
                    row=r + 12, column=c).value = req_data['download'][i]['pageLanguage']
            if 'pageAccessibilty' in req_data['download'][i]:
                ws.cell(
                    row=r + 13, column=c).value = req_data['download'][i]['pageAccessibilty']

            if 'userAuthState' in req_data['download'][i]:
                ws.cell(
                    row=r + 14, column=c).value = req_data['download'][i]['userAuthState']
            if 'userType' in req_data['download'][i]:
                ws.cell(
                    row=r + 15, column=c).value = req_data['download'][i]['userType']
            if 'userid' in req_data['download'][i]:
                ws.cell(
                    row=r + 16, column=c).value = req_data['download'][i]['userid']

            if 'eventError' in req_data['download'][i]:
                ws.cell(
                    row=r + 44, column=c).value = req_data['download'][i]['eventError']

            if 'events_download' in req_data['download'][i]:
                ws.cell(
                    row=r + 50, column=c).value = req_data['download'][i]['events_download']
            if 'download_filename' in req_data['download'][i]:
                ws.cell(
                    row=r + 51, column=c).value = req_data['download'][i]['download_filename']
            if 'download_filetype' in req_data['download'][i]:
                ws.cell(
                    row=r + 52, column=c).value = req_data['download'][i]['download_filetype']
            c = c + 1
        if req_data['MobileJson'] == 'True':
            self.json_file(data, wb)
        if req_data['TestCase'] == 'True':
            self.testCaseVRG(data,wb)
            self.testCaseMobileJSON(data,wb)
        wb.save(fname + '.xltx')
        return print('success')

    def json_file(self, req_data, wb):
        pages = {}
        for i in range(0, len(req_data['pages'])):
            if 'applicationType' in req_data:
                if 'pageName' in req_data['pages'][i]:
                    pages['state_' + req_data['pages'][i]
                    ['pageName'] + '_screen'] = {}
                    pages['state_' + req_data['pages'][i]
                    ['pageName'] + '_screen']['page'] = {}
                    pages['state_' + req_data['pages'][i]['pageName'] + '_screen']['page']['name'] = req_data['pages'][i][
                        'pageName']
                    pages['state_' + req_data['pages'][i]['pageName'] + '_screen']['page']['hierarchy'] = \
                        str(req_data['pages'][i]['pageHierarchy']).replace("'","").replace("[","").replace("]","").replace(" ","-").replace(".",">")
        interaction = {}
        for i in range(0, len(req_data['interaction'])):
            if 'applicationType' in req_data:
                if 'interactionName' in req_data['interaction'][i]:
                    interaction['action_' + str(req_data['interaction'][i]
                    ['interactionName']).lower().replace(" ","-").replace(":","-") + ''] = {}
                    interaction['action_' + str(req_data['interaction'][i]
                    ['interactionName']).lower().replace(" ","-").replace(":","-") + '']['interaction'] = {}
                    interaction['action_' + str(req_data['interaction'][i]['interactionName']).lower().replace(" ","-").replace(":","-") + '']['interaction'][
                        'name'] = str(req_data['interaction'][i]
                    ['interactionName']).lower().replace(" ","-")

        forms = {}
        for parent in range(0, len(req_data['forms'])):
            for child in range(0, len(req_data['forms'][parent]['steps'])):
                if 'pageName' in req_data['forms'][parent]['steps'][child]:
                    forms['state_' + req_data['forms'][parent]['steps']
                    [child]['pageName'] + '_screen'] = {}

                    # Product Object Created in Form JSON Object
                    if req_data['forms'][parent]['isTransactionExist']:
                        forms['state_' + req_data['forms'][parent]['steps']
                        [child]['pageName'] + '_screen']['product'] = [{}]
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['product'][0]['id'] = \
                           req_data['forms'][parent]['steps'][child]['productId']
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['product'][0]['positioning'] = \
                           req_data['forms'][parent]['steps'][child]['productPositioning']
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['product'][0]['grouping'] = \
                            req_data['forms'][parent]['steps'][child]['productGrouping']
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['product'][0]['parentproduct'] = \
                            req_data['forms'][parent]['steps'][child]['parentProduct']
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['product'][0]['adjudication'] = \
                            req_data['forms'][parent]['steps'][child]['adjudication']
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['product'][0]['fulfilment'] = \
                            req_data['forms'][parent]['steps'][child]['fulfillment']
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['product'][0]['paperless'] = \
                            req_data['forms'][parent]['steps'][child]['isPaperless']

                    # Page Object Created in Form JSON Object
                    forms['state_' + req_data['forms'][parent]['steps']
                    [child]['pageName'] + '_screen']['page'] = {}
                    forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['page']['hierarchy'] = \
                        str(req_data['forms'][parent]['steps'][child]['pageHierarchy']).replace("\\"," ").replace("[","").replace("]","").replace("'","").replace(",",">")
                    if req_data['forms'][parent]['steps'][child]['stepName'] != '':
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['page']['name'] = \
                            str(req_data['forms'][parent]['steps'][child]['stepName']).lower().replace(" ","-")
                    else:
                        arr = str(req_data['forms'][parent]['steps'][child]['pageHierarchy']).replace("\\"," ").replace("[","").replace("]","").replace("'","").split(" ")
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['page']['name']= \
                        str(arr[len(arr)-1]).lower().replace(" ","-")

                    # Form Object Created in Form JSON Object
                    forms['state_' + req_data['forms'][parent]['steps']
                    [child]['pageName'] + '_screen']['forms'] = {}
                    forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['forms']['name'] = \
                        str(req_data['forms'][parent]['formName']).lower().replace(" ","-")
                    if req_data['forms'][parent]['steps'][child]['stepName'] != '':
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['forms']['stepName'] = \
                        str(req_data['forms'][parent]['steps'][child]['stepName']).replace(" ","-").replace(".",">")
                    else:
                        arr = str(req_data['forms'][parent]['steps'][child]['pageHierarchy']).replace("\\"," ").replace("[","").replace("]","").replace("'","").split(" ")
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['forms']['stepName'] = \
                        str(arr[len(arr)-1]).lower().replace(" ","-")

                    # Event Object created in Form JSON Object
                    forms['state_' + req_data['forms'][parent]['steps']
                    [child]['pageName'] + '_screen']['events'] = {}
                    if (req_data['forms'][parent]['steps'][child]['formStep'] == 'true'):
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['events'][
                            'formStep'] = \
                            req_data['forms'][parent]['steps'][child]['formStep']
                    if (req_data['forms'][parent]['steps'][child]['formView'] == 'true'):
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['events'][
                            'formView'] = \
                            req_data['forms'][parent]['steps'][child]['formView']
                    if req_data['forms'][parent]['steps'][child]['formQualify'] == 'true':
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['events'][
                            'formQualify'] = \
                            req_data['forms'][parent]['steps'][child]['formQualify']
                    if req_data['forms'][parent]['steps'][child]['formsubmit'] == 'true':
                        forms['state_' + req_data['forms'][parent]['steps'][child]['pageName'] + '_screen']['events'][
                            'formsubmit'] = \
                            req_data['forms'][parent]['steps'][child]['formsubmit']

        dictMerge = {**forms, **pages, **interaction}
        print(json.dumps(dictMerge))
        with open('JsonData.json', 'w') as outfile:
            json.dump(dictMerge, outfile)
        ws1 = wb.create_sheet("JSON File", 1)
        ws1.cell(row=1, column=1).value = json.dumps(dictMerge)
        ws1.row_dimensions[0].height = 300
        ws1.column_dimensions['B'].width = 300
        ws1.cell(row=1, column=1).font = Font(size=12, name='Times New Roman')
        print(json.dumps(dictMerge))

    def testCaseVRG(self, req_data,workbook):

        alignment = Alignment(horizontal='left',vertical='top',wrapText=True)
        now = date.today()
        full = str(now.month) + "/" + str(now.day) + "/" + str(now.year)

        ws1 = workbook.create_sheet("VRG Test Cases",2)
        ws1.row_dimensions[1].height = 34.5

        ws1.cell(row = 1 , column= 1).value = 'Sr. No.'
        ws1.cell(row = 1 , column= 1).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
        ws1.cell(row = 1 , column= 2).value = 'Date Created'
        ws1.cell(row = 1 , column= 3).value = 'Subject'
        ws1.cell(row = 1 , column= 4).value = 'Test Case Name'
        ws1.cell(row = 1 , column= 5).value = 'Test Objective'
        ws1.cell(row = 1 , column= 6).value = 'Step Name'
        ws1.cell(row = 1 , column= 7).value = 'Description'
        ws1.cell(row = 1 , column= 8).value = 'Expected Result'
        ws1.cell(row = 1 , column= 9).value = 'Author'
        for i in range(1,10):
             ws1.cell(row = 1 , column= i).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
             ws1.cell(row = 1 , column= i).font = Font(bold=True,color='008000')

        ws1.column_dimensions['B'].width = 10
        ws1.column_dimensions['D'].width = 30
        ws1.column_dimensions['F'].width = 30
        ws1.column_dimensions['E'].width = 14.89
        ws1.column_dimensions['G'].width = 25
        ws1.column_dimensions['H'].width = 120

        id = 1
        r = 2
        c = 1
        s = 'Step'
        e = 'To Verify Adobe variables match those detailed in the VRG \n \n Pre Condition: Debugger tool to be launched for testing purpose(Recommended  Adobe Pulse Debugger)'
        g = 'Ensure Adobe Debugger is active'

        for i in range(0, len(req_data['pages'])):
            event = 'Events: '
            comment = ' //'
            if 'pageName' in req_data['pages'][i]:
                ws1.cell(row=r, column=c).value = str(id)
                ws1.cell(row = r , column= c).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row=r, column=c + 1).value = full
                ws1.cell(row=r, column=c + 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

                ws1.cell(row=r, column=c + 2).value = req_data['ProjectName']
                ws1.cell(row=r, column=c + 2).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 4).value = e
                ws1.cell(row=r, column=c + 4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)

                ws1.cell(row=r, column=c + 6).value = g
                ws1.cell(row=r, column=c + 6).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

                val = req_data['pages'][i]['pageName']
                b = '-.'
                for char in b:
                    val = val.replace(char, " ")
                ws1.cell(row=r, column=c + 3).value = val.title()
                ws1.cell(row=r, column=c + 5).value = req_data['pages'][i]['pageName']

                ws1.cell(row=r, column=c + 3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 5).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)

            if 'sitebrand' in req_data:
                if (bool(req_data['pages'][i]['loginEvent'])):
                    event = event + ' event19 ,'
                    comment = comment + 'Login Event, '
                if (bool(req_data['pages'][i]['eventError'])):
                    event = event + 'event3'
                    comment = comment + 'Error Event '
                dct = 'pageName, eVar10: ' + str(
                    req_data['sitebrand']) + '>' + str(req_data['siteName']) + '>' + str(
                    req_data['pages'][i]['pageName']) + '    //Unique for each Page' + '\n eVar11: ' + str(
                    req_data['siteAppVersion']) + ':' + str(req_data['siteLastBuildDate']) + ':' + str(
                    req_data['siteName']) + ':' + str(req_data[
                                                          'siteType']) + '    //siteAppVersion,siteLastBuildDate,siteName,siteType' + '\neVar15: ' + str(
                    req_data['pages'][i]['pageUrl']) + '    //Page URL' + '\n eVar9: ' + str(
                    req_data['pages'][i]['pageReferrer']) + '    //Previous Page Name' + '\n eVar14: ' + str(
                    req_data['pages'][i]['pageHierarchy']) + '    // Hierarchy of the Page' + '\n eVar16: ' + str(
                    req_data['pages'][i]['pageLanguage']) + '    // Language Supported For Page' + '\n prop12: ' + str(
                    req_data['pages'][i]['pageAccessibilty']) + '\n eVar19: {' + str(
                    req_data['pages'][i]['userAuthState']) + ' : ' + str(
                    req_data['pages'][i][
                        'userType']) + '}' + '    // Authentication of User and User Type' + '\n eVar20: ' + str(
                    req_data['pages'][i]['userId']) + '    // User ID' + '\n \n' + event + '\t' + comment
                ws1.cell(row=r, column=c + 7).value = str(dct)
                ws1.cell(row=r, column=c + 7).alignment = alignment
                ws1.cell(row=r, column=c + 8).value = req_data['userName']
                ws1.cell(row = r , column= c + 8).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.row_dimensions[r].height = 180
            id = id + 1
            r = r + 1

        for i in range(0, len(req_data['interaction'])):
            event = 'Events: '
            comment = ' '
            if 'pageName' in req_data['interaction'][i]:
                ws1.cell(row=r, column=c).value = str(id)
                ws1.cell(row = r , column= c ).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                ws1.cell(row=r, column=c + 1).value = full
                ws1.cell(row = r , column= c + 1).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                ws1.cell(row=r, column=c + 2).value = req_data['ProjectName']
                ws1.cell(row = r , column= c + 2).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 4).value = e
                ws1.cell(row = r , column= c + 4).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='top',wrapText=True)

                ws1.cell(row=r, column=c +6).value = g
                ws1.cell(row = r , column= c + 6).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                val = req_data['interaction'][i]['interactionName']
                b = '-.'
                for char in b:
                    val = val.replace(char, " ")
                ws1.cell(row=r, column=c + 3).value = val.title()
                ws1.cell(row = r , column= c + 3).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 5).value = req_data['interaction'][i]['interactionName']
                ws1.cell(row = r , column= c + 5).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)

            if 'sitebrand' in req_data:
                if (bool(req_data['interaction'][i]['siteInteractionEvent'])):
                    event = event + ' event24 ,'
                    comment = comment + '// Site Interaction Event, '
                dct = 'pageName, eVar10: ' + str(
                    req_data['sitebrand']) + '>' + str(req_data['siteName']) + '>' + str(
                    req_data['interaction'][i]['pageName']) + '    //Unique for each Page' + '\n eVar11: ' + str(
                    req_data['siteAppVersion']) + ':' + str(req_data['siteLastBuildDate']) + ':' + str(
                    req_data['siteName']) + ':' + str(req_data[
                                                          'siteType']) + '    //siteAppVersion,siteLastBuildDate,siteName,siteType' + '\n eVar15: ' + str(
                    req_data['interaction'][i]['pageUrl']) + '    //Page URL' + '\n eVar9: ' + str(
                    req_data['interaction'][i]['pageReferrer']) + '    //Previous Page Name' + '\n eVar14: ' + str(
                    req_data['interaction'][i]['pageHierarchy']) + '    // Hierarchy of the Page' + '\n eVar16: ' + str(
                    req_data['interaction'][i][
                        'pageLanguage']) + '    // Language Supported For Page' + '\n prop12: ' + str(
                    req_data['interaction'][i]['pageAccessibilty']) + '' + '\n eVar24: ' + str(
                    req_data['interaction'][i]['interactionName']) + " //Interaction Name \n" + event + '\t' + comment
                ws1.cell(row=r, column=c + 7).value = str(dct)
                ws1.cell(row=r, column=c + 7).alignment= alignment
                ws1.cell(row=r, column=c + 8).value = req_data['userName']
                ws1.cell(row = r , column= c + 8).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.row_dimensions[r].height = 180
            id = id + 1
            r = r + 1

        for i in range(0, len(req_data['download'])):
            event = ''
            comment = ''
            if 'pageName' in req_data['download'][i]:
                ws1.cell(row=r, column=c).value = str(id)
                ws1.cell(row = r , column= c).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                ws1.cell(row=r, column=c + 1).value = full
                ws1.cell(row = r , column= c + 1).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                ws1.cell(row=r, column=c + 2).value = req_data['ProjectName']
                ws1.cell(row = r , column= c + 2).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 4).value = e
                ws1.cell(row = r , column= c + 4).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='top',wrapText=True)

                ws1.cell(row=r, column=c + 6).value = g
                ws1.cell(row = r , column= c + 6).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                val = req_data['download'][i]['download_filename']
                b = '-.'
                for char in b:
                    val = val.replace(char, " ")
                ws1.cell(row=r, column=c + 3).value = val.title()
                ws1.cell(row = r , column= c + 3).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)

                ws1.cell(row=r, column=c + 5).value = req_data['download'][i]['download_filename']
                ws1.cell(row = r , column= c + 5).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)

            if 'sitebrand' in req_data:
                if (bool(req_data['download'][i]['events_download'])):
                    event = event + ' event21 ,'
                    comment = comment + '// Download Event, '
                dct = 'pageName, eVar10: ' + str(
                    req_data['sitebrand']) + '>' + str(req_data['siteName']) + '>' + str(
                    req_data['download'][i]['pageName']) + '    //Unique for each Page' + '\n eVar11: ' + str(
                    req_data['siteAppVersion']) + ':' + str(req_data['siteLastBuildDate']) + ':' + str(
                    req_data['siteName']) + ':' + str(req_data[
                                                          'siteType']) + '    //siteAppVersion,siteLastBuildDate,siteName,siteType' + '\n eVar15: ' + str(
                    req_data['download'][i]['pageUrl']) + '    //Page URL' + '\n eVar9: ' + str(
                    req_data['download'][i]['pageReferrer']) + '    //Previous Page Name' + '\n eVar14: ' + str(
                    req_data['download'][i]['pageHierarchy']) + '    // Hierarchy of the Page' + '\n eVar16: ' + str(
                    req_data['download'][i][
                        'pageLanguage']) + '    // Language Supported For Page' + '\n prop12: ' + str(
                    req_data['download'][i]['pageAccessibilty']) + '\n eVar19: {' + str(
                    req_data['download'][i]['userAuthState']) + ' : ' + str(
                    req_data['download'][i][
                        'userType']) + '}' + '    // Authentication of User and User Type' + '\n eVar20: ' + str(
                    req_data['download'][i]['userId']) + '    // User ID' + '\n eVar24: ' + str(
                    req_data['download'][i]['download_filename']) + ':' + str(req_data['download'][i][
                                                                                  'download_filetype']) + '    // Download FileName:FileType \n' + event + '\t' + comment
                ws1.cell(row=r, column=c + 7).value = str(dct)
                ws1.cell(row = r , column= c + 7).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='top',wrapText=True)

                ws1.cell(row=r, column=c + 8).value = req_data['userName']
                ws1.cell(row = r , column= c + 8).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                ws1.row_dimensions[r].height = 180
            id = id + 1
            r = r + 1

        for parent in range(0, len(req_data['forms'])):
            for child in range(0, len(req_data['forms'][parent]['steps'])):
                event = 'Events'
                comment = ''
                if (bool(req_data['forms'][parent]['steps'][child]['formStep'])):
                    event = event + 'event4'
                if (bool(req_data['forms'][parent]['steps'][child]['formView'])):
                    event = event + ', event1'
                if (bool(req_data['forms'][parent]['steps'][child]['formsubmit'])):
                    event = event + ', event2'
                if (bool(req_data['forms'][parent]['steps'][child]['formQualify'])):
                    event = event + ', event5'
                if (bool(req_data['forms'][parent]['steps'][child]['isExternal'])):
                    event = event + ', event46'
                if (bool(req_data['forms'][parent]['steps'][child]['pageLogin'])):
                    event = event + ', event19'
                if (bool(req_data['forms'][parent]['steps'][child]['eventError'])):
                    event = event + ', event3'
                if (bool(req_data['forms'][parent]['steps'][child]['isPaperless'])):
                    event = event + ', event30'
                if (bool(req_data['forms'][parent]['steps'][child]['productRecommendation'])):
                    event = event + ', event103'

                if 'pageName' in req_data['forms'][parent]['steps'][child]:
                    ws1.cell(row=r, column=c).value = req_data['forms'][parent]['steps'][child]['pageName']
                    ws1.cell(row = r , column= c).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                    ws1.cell(row=r, column=1).value = str(id)
                    ws1.cell(row = r , column= 1).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                    ws1.cell(row=r, column=2).value = full
                    ws1.cell(row = r , column= 2).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                    ws1.cell(row=r, column=2).font = openpyxl.styles.Font(bold=True)

                    ws1.cell(row=r, column=3).value = req_data['ProjectName']
                    ws1.cell(row = r , column= 3).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                    ws1.cell(row=r, column=3).font = openpyxl.styles.Font(bold=True)

                    ws1.cell(row=r, column=5).value = e
                    ws1.cell(row = r , column= 5).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                    ws1.cell(row=r, column=7).value = g
                    ws1.cell(row = r , column= 7).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                    val = req_data['forms'][parent]['steps'][child]['pageName']
                    b = '-.'
                    for char in b:
                        val = val.replace(char, " ")

                    ws1.cell(row=r, column=4).value = val.title()
                    ws1.cell(row = r , column= 4).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                    ws1.cell(row=r, column=6).value = val.title()
                    ws1.cell(row = r , column= 6).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)

                    dt = 'pageName, eVar10: ' + str(
                        req_data['sitebrand']) + '>' + str(req_data['siteName']) + '>' + str(
                        req_data['forms'][parent]['steps'][child][
                            'pageName']) + '    //Unique for each Page \n eVar11:' + str(
                        req_data['siteAppVersion']) + ':' + str(req_data['siteLastBuildDate']) + ':' + str(
                        req_data['siteName']) + ':' + str(req_data[
                                                              'siteType']) + '    //siteAppVersion:siteLastBuildDate:siteName:siteType' + '\n eVar15: ' + str(
                        req_data['forms'][parent]['steps'][child]['pageUrl']) + '    //Page URL' + '\n eVar9: ' + str(
                        req_data['forms'][parent]['steps'][child][
                            'pageReferrer']) + ' //Previous Page Name' + '\n eVar14: ' + str(
                        req_data['forms'][parent]['steps'][child][
                            'pageHierarchy']) + ' // Hierarchy of the Page' + '\n eVar16: ' + str(
                        req_data['forms'][parent]['steps'][child][
                            'pageLanguage']) + '    // Language Supported For Page' + '\n prop12: ' + str(
                        req_data['forms'][parent]['steps'][child]['pageAccessibilty']) + '\n eVar19: {' + str(
                        req_data['forms'][parent]['steps'][child]['userAuthState']) + ' : ' + str(
                        req_data['forms'][parent]['steps'][child][
                            'userType']) + '}' + ' // Authentication of User and User Type ' + '\n eVar20: ' + str(
                        req_data['forms'][parent]['steps'][child]['userId']) + ' // User ID' + '\n evar1: ' + str(
                        req_data['forms'][parent]['formName']) + ' // Form Name' + '\n eVar4: ' + str(
                        req_data['forms'][parent]['steps'][child][
                            'stepName']) + ' //Form Step Name' + ' \n eVar5: ' + str(
                        req_data['forms'][parent]['steps'][child]['uniqueId']) + ' // Unique Id of Form'
                    if (req_data['forms'][parent]['isTransactionExist']):
                        dt = dt + '\n eVar41: ' + str(
                            req_data['forms'][parent]['steps'][child]['fromtransaction']) + ':' + str(
                            req_data['forms'][parent]['steps'][child][
                                'totransaction']) + '    //Transaction From: Transaction To'
                    if (req_data['forms'][parent]['isProductExist']):
                        dt = dt + '\n eVar82: ' + str(req_data['forms'][parent]['steps'][child][
                                                          'productPositioning']) + '    //Product Positioning \n eVar83 : ' + str(
                            req_data['forms'][parent]['steps'][child][
                                'productGrouping']) + '    //Product grouping \n eVar84: ' + str(
                            req_data['forms'][parent]['steps'][child][
                                'parentProduct']) + ' // Parent Product \neVar88:' + str(
                            req_data['forms'][parent]['steps'][child]['adjudication']) + '   // Adjudication\n ' + event
                    ws1.cell(row=r, column=c + 7).value = str(dt)
                    ws1.cell(row = r , column= c+7).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='top',wrapText=True)
                    ws1.cell(row=r, column=c + 8).value = req_data['userName']
                    ws1.cell(row = r , column= c + 8).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                    ws1.row_dimensions[r].height = 203
                    #ws1.row_dimensions[r].height = 180
                    id = id + 1
                    r = r + 1

    def testCaseMobileJSON(self,req_data,workbook):
        border = Border(left=Side(border_style='thin',
                         color='FF000000'),
                 right=Side(border_style='thin',
                            color='FF000000'),
                 top=Side(border_style='thin',
                          color='FF000000'),
                 bottom=Side(border_style='thin',
                             color='FF000000'),
                )

        e = 'To Verify Adobe variables match those detailed in the JSON. \n \n Pre Condition: Charles Proxy tool to be launched for testing purpose.'
        g = 'Ensure Charles Proxy is active'

        note = 'Please note: All native requests are POST requests \n Check Request Headers in Charles Proxy'
        developement = 'Dev '
        production = 'Prod '
        message='environment: All requests are for "POST /b/ss/cibcnativeappdev/0/'
        code='xxxxxx'

        ws1 = workbook.create_sheet("Mobile JSON VRG",3)
        ws1.row_dimensions[1].height = 24
        ws1.row_dimensions[2].height = 63.6

        # set the width of column in Mobile JSON Test Sheet
        ws1.column_dimensions['B'].width = 10
        ws1.column_dimensions['D'].width = 30
        ws1.column_dimensions['F'].width = 30
        ws1.column_dimensions['E'].width = 14.89
        ws1.column_dimensions['G'].width = 25
        ws1.column_dimensions['H'].width = 120

        #Column Header in Mobile JSON Test Case Sheet
        ws1.cell(row = 1 , column= 1).value = 'Sr. No.'

        ws1.cell(row = 1 , column= 2).value = 'Date Created'
        ws1.cell(row = 1 , column= 3).value = 'Subject'
        ws1.cell(row = 1 , column= 4).value = 'Test Case Name'
        ws1.cell(row = 1 , column= 5).value = 'Test Objective'
        ws1.cell(row = 1 , column= 6).value = 'Step Name'
        ws1.cell(row = 1 , column= 7).value = 'Description'
        ws1.cell(row = 1 , column= 8).value = 'Expected Result'
        ws1.cell(row = 1 , column= 9).value = 'Author'

        for i in range(1,10):
             ws1.cell(row = 1 , column= i).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
             ws1.cell(row = 1 , column= i).font = Font(bold=True,color='008000')
             ws1.cell(row = 1 , column= i).border = border

        ws1.cell(row = 2, column= 1).value = note + '\n \t'+developement+message+code+'\n \t'+production+message+code
        ws1.cell(row = 2 , column= 1).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='top',wrapText=True)
        ws1.cell(row = 2 , column= 1).font = Font(color='008000')
        ws1.cell(row = 2 , column= 1).border = border

        ws1.merge_cells('A2:G2')
        ws1.merge_cells('H2:Z2')

        id = 1
        r = 3
        c = 1
        now = date.today()
        full = str(now.month) + "/" + str(now.day) + "/" + str(now.year)

        for i in range(0, len(req_data['pages'])):
            event = 'Events: '
            comment = ' //'
            ws1.row_dimensions[r].height = 203.6
            if 'pageName' in req_data['pages'][i]:
                ws1.cell(row=r, column=c).value = str(id)
                ws1.cell(row = r , column= c).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row = r , column= c).border = border

                ws1.cell(row=r, column=c + 1).value = full
                ws1.cell(row=r, column=c + 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1.cell(row = r , column= c + 1).border = border

                ws1.cell(row=r, column=c + 2).value = req_data['ProjectName']
                ws1.cell(row=r, column=c + 2).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row = r , column= c + 2).border = border

                ws1.cell(row=r, column=c + 4).value = e
                ws1.cell(row=r, column=c + 4).alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
                ws1.cell(row = r , column= c + 4).border = border

                ws1.cell(row=r, column=c + 6).value = g
                ws1.cell(row=r, column=c + 6).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1.cell(row = r , column= c + 6).border = border

                val = req_data['pages'][i]['pageName']
                b = '-.'
                for char in b:
                    val = val.replace(char, " ")
                ws1.cell(row=r, column=c + 3).value = val.title()
                ws1.cell(row=r, column=c + 5).value = req_data['pages'][i]['pageName']

                ws1.cell(row=r, column=c + 3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row = r , column= c + 3).border = border
                ws1.cell(row=r, column=c + 5).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row = r , column= c + 5).border = border
            if 'sitebrand' in req_data:
                dct = 'pageName: ' + str(
                    req_data['sitebrand']) + '>' + str(req_data['siteName']) + '>' + str(
                    req_data['pages'][i]['pageName']) + '  //Unique for each Page' + '\n '+'site.name' + str(
                    req_data['siteName']) + '  //Name of the Site \n site.type : ' + str(req_data['siteType']) + '  //Type of the Site' + '\n page.url:' + str(
                    req_data['pages'][i]['pageUrl']) + '  // URL of the Page' + '\n page.referrer: ' + str(
                    req_data['pages'][i]['pageReferrer']) + '  //Previous Page Name' + '\n page.hierarchy: ' + str(
                    req_data['pages'][i]['pageHierarchy']) + ' // Hierarchy of the Page' + '\n page.language: ' + str(
                    req_data['pages'][i]['pageLanguage']) + '    // Language Supported For Page' + '\n page.accessibilty: ' + str(
                    req_data['pages'][i]['pageAccessibilty']) + ' //Page Accessibility \n user.Authstate:' + str(
                    req_data['pages'][i]['userAuthState']) + ' // depending on pre sign on or post sign on \n user.Type' + str(
                    req_data['pages'][i]['userType'])+ ' //Type of the User \n user.ID: ' + '' + str(
                    req_data['pages'][i]['userId']) + '   //user Id for the post sign on. it should be consistent throughout the whole app. ' + '\n \n'
                ws1.cell(row=r, column=c + 7).value = str(dct)
                ws1.cell(row=r, column=c + 7).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='top',wrapText=True)
                ws1.cell(row = r , column= c + 7).border = border

                ws1.cell(row=r, column=c + 8).value = req_data['userName']
                ws1.cell(row = r , column= c + 8).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row = r , column= c + 8).border = border
                ws1.row_dimensions[r].height = 180
            id = id + 1
            r = r + 1
        for i in range(0, len(req_data['interaction'])):
            #ws1.row_dimensions[r].height = 203.6
            if 'pageName' in req_data['interaction'][i]:

                ws1.cell(row=r, column=c).value = str(id)
                ws1.cell(row = r , column= c ).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row = r , column= c + 2).border = border

                ws1.cell(row=r, column=c + 1).value = full
                ws1.cell(row = r , column= c+ 1).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row = r , column= c + 1).border = border

                ws1.cell(row=r, column=c + 2).value = req_data['ProjectName']
                ws1.cell(row = r , column= c + 2).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row = r , column= c + 2).border = border

                ws1.cell(row=r, column=c + 4).value = e
                ws1.cell(row = r , column= c+4).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='top',wrapText=True)
                ws1.cell(row = r , column= c + 4).border = border

                ws1.cell(row=r, column=c +6).value = g
                ws1.cell(row = r , column= c+6).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row = r , column= c + 6).border = border

                val = req_data['interaction'][i]['interactionName']
                b = '-.'
                for char in b:
                    val = val.replace(char, " ")
                ws1.cell(row=r, column=c + 3).value = val.title()
                ws1.cell(row = r , column= c+3).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row = r , column= c + 3).border = border

                ws1.cell(row=r, column=c + 5).value = req_data['interaction'][i]['interactionName']
                ws1.cell(row = r , column= c+5).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row = r , column= c + 5).border = border

            if 'sitebrand' in req_data:
                dct = 'pageName' + str(
                    req_data['sitebrand']) + '>' + str(req_data['siteName']) + '>' + str(
                    req_data['interaction'][i]['pageName']) + '    //Unique for each Page {Interaction Page}' + '\n'+ 'site.name :' + str(
                    req_data['siteName']) + ' //Name of the Site\n site.type:' + str(req_data['siteType']) + '  //Type of the site' + '\n page.url: ' + str(
                    req_data['interaction'][i]['pageUrl']) + '//URL of the Page' + '\n page.referreer:' + str(
                    req_data['interaction'][i]['pageReferrer']) + '//Previous Page Name' + '\n page.hierarchy:' + str(
                    req_data['interaction'][i]['pageHierarchy']) + '  // Hierarchy of the Page' + '\n page.language: ' + str(
                    req_data['interaction'][i][
                        'pageLanguage']) + '    // Language Supported For Page' + '\n page.accessibility:' + str(
                    req_data['interaction'][i]['pageAccessibilty']) + '//Page Accessibility ' + '\n interaction.name: ' + str(
                    req_data['interaction'][i]['interactionName']) + " //Interaction Name \n event.siteInteraction: " + req_data['interaction'][i]['siteInteractionEvent'] + '// Site Interaction Event '
                ws1.cell(row=r, column=c + 7).value = str(dct)
                ws1.cell(row = r , column= c + 7).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='top',wrapText=True)
                ws1.cell(row = r , column= c + 7).border = border

                ws1.cell(row=r, column=c + 8).value = req_data['userName']
                ws1.cell(row = r , column= c + 8).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                ws1.cell(row = r , column= c + 8).border = border

                ws1.row_dimensions[r].height = 203
            id = id + 1
            r = r + 1
        for parent in range(0, len(req_data['forms'])):
            for child in range(0, len(req_data['forms'][parent]['steps'])):
                if 'pageName' in req_data['forms'][parent]['steps'][child]:
                    ws1.cell(row=r, column=c).value = req_data['forms'][parent]['steps'][child]['pageName']
                    ws1.cell(row = r , column= c).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                    ws1.cell(row=r, column=1).value = str(id)
                    ws1.cell(row = r , column= 1).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                    ws1.cell(row = r , column= 1).border = border

                    ws1.cell(row=r, column=2).value = full
                    ws1.cell(row = r , column= 2).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                    ws1.cell(row=r, column=c + 2).font = openpyxl.styles.Font(bold=True)
                    ws1.cell(row = r , column= c + 2).border = border

                    ws1.cell(row=r, column=3).value = req_data['ProjectName']
                    ws1.cell(row = r , column= 3).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                    ws1.cell(row=r, column=c + 3).font = openpyxl.styles.Font(bold=True)
                    ws1.cell(row = r , column= c + 3).border = border

                    ws1.cell(row=r, column=5).value = e
                    ws1.cell(row = r , column= 5).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='top',wrapText=True)
                    ws1.cell(row=r, column=c + 5).font = openpyxl.styles.Font(bold=True)
                    ws1.cell(row = r , column= c + 5).border = border

                    ws1.cell(row=r, column=7).value = g
                    ws1.cell(row = r , column= 7).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                    ws1.cell(row = r , column= c + 7).border = border

                    val = req_data['forms'][parent]['steps'][child]['pageName']
                    b = '-.'
                    for char in b:
                        val = val.replace(char, " ")

                    ws1.cell(row=r, column=4).value = val.title()
                    ws1.cell(row = r , column= 4).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                    ws1.cell(row = r , column= c + 4).border = border

                    ws1.cell(row=r, column=6).value = val.title()
                    ws1.cell(row = r , column= 6).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                    ws1.cell(row = r , column= c + 6).border = border

                    dt = 'pageName:' + str(
                        req_data['sitebrand']) + '>' + str(req_data['siteName']) + '>' + str(
                        req_data['forms'][parent]['steps'][child][
                            'pageName']) + ' //Unique for each Page \n '+ 'site.name:' + str(
                        req_data['siteName']) + ' //Site Name \n site.type:' + str(req_data[
                                                              'siteType']) + '    //Type of the Site' + '\n page.url: ' + str(
                        req_data['forms'][parent]['steps'][child]['pageUrl']) + '  //URL of the Page' + '\n page.referrer: ' + str(
                        req_data['forms'][parent]['steps'][child][
                            'pageReferrer']) + ' //Previous Page Name' + '\n page.hierarchy' + str(
                        req_data['forms'][parent]['steps'][child][
                            'pageHierarchy']) + ' // Hierarchy of the Page' + '\n page.language: ' + str(
                        req_data['forms'][parent]['steps'][child][
                            'pageLanguage']) + '// Language Supported For Page' + '\n page.accessibility: ' + str(
                        req_data['forms'][parent]['steps'][child]['pageAccessibilty']) + '//Page Accessibility \n user.Authstate:' + str(
                        req_data['forms'][parent]['steps'][child]['userAuthState']) + '// depending on pre sign on or post sign on \n user.Type' + str(
                        req_data['forms'][parent]['steps'][child][
                            'userType']) + '}' + ' //Type of the User \n user.ID: ' + '' + str(
                        req_data['forms'][parent]['steps'][child]['userId']) + ' //user Id for the post sign on. it should be consistent throughout the whole app. ' + '\n form.name: ' + str(
                        req_data['forms'][parent]['formName']) + ' // Form Name' + '\n step.name: ' + str(
                        req_data['forms'][parent]['steps'][child][
                            'stepName']) + ' //Form Step Name' + ' \n unique.id: ' + str(
                        req_data['forms'][parent]['steps'][child]['uniqueId']) + ' // Unique Id of Form'
                    if req_data['forms'][parent]['isTransactionExist'] == 'true':
                        dt = dt + '\n transaction.from: ' + str(
                            req_data['forms'][parent]['steps'][child]['fromtransaction']) + '// Transaction From \n transaction.to: ' + str(
                            req_data['forms'][parent]['steps'][child][
                                'totransaction']) + ' //To Transaction'
                    if req_data['forms'][parent]['isProductExist'] == 'true':
                        dt = dt + '\n positioning: ' + str(req_data['forms'][parent]['steps'][child][
                                                          'productPositioning']) + '// Product Positioning\n grouping: ' + str(
                            req_data['forms'][parent]['steps'][child][
                                'productGrouping']) + '    //Product grouping \n parentProduct: ' + str(
                            req_data['forms'][parent]['steps'][child][
                                'parentProduct']) + ' // Parent Product \nadjudication:' + str(
                            req_data['forms'][parent]['steps'][child]['adjudication']) + '   // Adjudication\n '


                    ws1.cell(row=r, column=c + 7).value = str(dt)
                    ws1.cell(row = r , column= c + 7).alignment = openpyxl.styles.Alignment(horizontal='left',vertical='top',wrapText=True)
                    ws1.cell(row = r , column= c + 7).border = border

                    ws1.cell(row=r, column=c + 8).value = req_data['userName']
                    ws1.cell(row = r , column= c + 8).alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center',wrapText=True)
                    ws1.cell(row = r , column= c + 8).border = border

                    ws1.row_dimensions[r].height = 203
                    id = id + 1
                    r = r + 1
